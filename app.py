import os
import csv
import sqlite3
from io import StringIO, BytesIO
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, make_response, flash)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-secret-key')

# 数据库文件路径（PythonAnywhere 和本地通用）
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'shipments.db')

# ==================== 数据库 ====================

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    # 供应商账号表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(100) NOT NULL,
            company_name VARCHAR(100) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 发货记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS shipments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
            project_name VARCHAR(200) NOT NULL,
            project_code VARCHAR(100),
            purchase_order VARCHAR(100),
            customer_order VARCHAR(100),
            logistics_no VARCHAR(200),
            logistics_receipt VARCHAR(200),
            ship_date DATE,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 管理员表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(100) NOT NULL
        )
    ''')
    conn.commit()
    cursor.close()
    conn.close()

# ==================== 装饰器 ====================

def supplier_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('supplier_id'):
            return redirect(url_for('supplier_login'))
        return f(*args, **kwargs)
    return decorated

def admin_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_id'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

# ==================== 供应商路由 ====================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/supplier/login', methods=['GET', 'POST'])
def supplier_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM suppliers WHERE username = ? AND password = ?', (username, password))
        supplier = cursor.fetchone()
        cursor.close()
        conn.close()
        if supplier:
            session['supplier_id'] = supplier['id']
            session['supplier_name'] = supplier['company_name']
            session['supplier_username'] = supplier['username']
            return redirect(url_for('supplier_dashboard'))
        return render_template('supplier_login.html', error=True)
    return render_template('supplier_login.html', error=False)

@app.route('/supplier/logout')
def supplier_logout():
    session.pop('supplier_id', None)
    session.pop('supplier_name', None)
    session.pop('supplier_username', None)
    return redirect(url_for('index'))

@app.route('/supplier/dashboard')
@supplier_login_required
def supplier_dashboard():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'SELECT * FROM shipments WHERE supplier_id = ? ORDER BY created_at DESC',
        (session['supplier_id'],)
    )
    shipments = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('supplier_dashboard.html', shipments=shipments)

@app.route('/supplier/submit', methods=['GET', 'POST'])
@supplier_login_required
def supplier_submit():
    if request.method == 'POST':
        project_name = request.form.get('project_name', '').strip()
        project_code = request.form.get('project_code', '').strip()
        purchase_order = request.form.get('purchase_order', '').strip()
        customer_order = request.form.get('customer_order', '').strip()
        logistics_no = request.form.get('logistics_no', '').strip()
        logistics_receipt = request.form.get('logistics_receipt', '').strip()
        ship_date = request.form.get('ship_date', '') or None
        remark = request.form.get('remark', '').strip()

        if not project_name:
            return render_template('supplier_submit.html', error='项目名称不能为空')

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO shipments
            (supplier_id, project_name, project_code, purchase_order,
             customer_order, logistics_no, logistics_receipt, ship_date, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (session['supplier_id'], project_name, project_code,
              purchase_order, customer_order, logistics_no,
              logistics_receipt, ship_date, remark))
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('supplier_dashboard'))
    return render_template('supplier_submit.html', error=None)

@app.route('/supplier/edit/<int:item_id>', methods=['GET', 'POST'])
@supplier_login_required
def supplier_edit(item_id):
    conn = get_db()
    cursor = conn.cursor()
    if request.method == 'POST':
        project_name = request.form.get('project_name', '').strip()
        project_code = request.form.get('project_code', '').strip()
        purchase_order = request.form.get('purchase_order', '').strip()
        customer_order = request.form.get('customer_order', '').strip()
        logistics_no = request.form.get('logistics_no', '').strip()
        logistics_receipt = request.form.get('logistics_receipt', '').strip()
        ship_date = request.form.get('ship_date', '') or None
        remark = request.form.get('remark', '').strip()

        cursor.execute('''
            UPDATE shipments SET
                project_name=?, project_code=?, purchase_order=?,
                customer_order=?, logistics_no=?, logistics_receipt=?,
                ship_date=?, remark=?
            WHERE id=? AND supplier_id=?
        ''', (project_name, project_code, purchase_order,
              customer_order, logistics_no, logistics_receipt,
              ship_date, remark, item_id, session['supplier_id']))
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('supplier_dashboard'))

    cursor.execute('SELECT * FROM shipments WHERE id=? AND supplier_id=?',
                   (item_id, session['supplier_id']))
    item = cursor.fetchone()
    cursor.close()
    conn.close()
    if not item:
        return redirect(url_for('supplier_dashboard'))
    return render_template('supplier_edit.html', item=item)

@app.route('/supplier/delete/<int:item_id>', methods=['POST'])
@supplier_login_required
def supplier_delete(item_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM shipments WHERE id=? AND supplier_id=?',
                   (item_id, session['supplier_id']))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('supplier_dashboard'))

@app.route('/supplier/export/excel')
@supplier_login_required
def supplier_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return supplier_export_csv()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'SELECT * FROM shipments WHERE supplier_id = ? ORDER BY created_at DESC',
        (session['supplier_id'],)
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '发货记录'

    headers = ['项目名称', '项目编号', '采购订单号', '客户订单号',
                '物流单号', '物流回单', '发货日期', '备注', '提交时间']
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='E60012', end_color='E60012', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        values = [row['project_name'], row['project_code'], row['purchase_order'],
                  row['customer_order'], row['logistics_no'], row['logistics_receipt'],
                  str(row['ship_date']) if row['ship_date'] else '',
                  row['remark'] or '',
                  str(row['created_at']) if row['created_at'] else '']
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                      for r in range(1, len(rows) + 2))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    supplier_name = session.get('supplier_name', 'supplier')
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={supplier_name}_发货记录.xlsx'
    return response

# ==================== 管理员路由 ====================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM admins WHERE username = ? AND password = ?',
                       (username, password))
        admin = cursor.fetchone()
        cursor.close()
        conn.close()
        if admin:
            session['admin_id'] = admin['id']
            session['admin_username'] = admin['username']
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_login.html', error=True)
    return render_template('admin_login.html', error=False)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    return redirect(url_for('index'))

@app.route('/admin')
@admin_login_required
def admin_dashboard():
    filter_supplier = request.args.get('supplier_id', '')
    keyword = request.args.get('keyword', '').strip()

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id, company_name FROM suppliers ORDER BY company_name')
    suppliers = cursor.fetchall()

    query = '''
        SELECT s.*, sp.company_name
        FROM shipments s
        JOIN suppliers sp ON s.supplier_id = sp.id
        WHERE 1=1
    '''
    params = []
    if filter_supplier:
        query += ' AND s.supplier_id = ?'
        params.append(int(filter_supplier))
    if keyword:
        query += ''' AND (s.project_name LIKE ?
                      OR s.project_code LIKE ?
                      OR s.purchase_order LIKE ?
                      OR s.customer_order LIKE ?
                      OR s.logistics_no LIKE ?)'''
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw, kw, kw])
    query += ' ORDER BY s.created_at DESC'

    cursor.execute(query, params)
    shipments = cursor.fetchall()

    cursor.execute('SELECT COUNT(*) FROM shipments')
    total_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(DISTINCT supplier_id) FROM shipments')
    supplier_count = cursor.fetchone()[0]

    cursor.close()
    conn.close()
    return render_template('admin_dashboard.html',
                           shipments=shipments, suppliers=suppliers,
                           filter_supplier=filter_supplier, keyword=keyword,
                           total_count=total_count, supplier_count=supplier_count)

@app.route('/admin/supplier/manage')
@admin_login_required
def admin_supplier_manage():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT sp.*, COUNT(s.id) as shipment_count
        FROM suppliers sp
        LEFT JOIN shipments s ON sp.id = s.supplier_id
        GROUP BY sp.id
        ORDER BY sp.company_name
    ''')
    suppliers = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('admin_supplier_manage.html', suppliers=suppliers)

@app.route('/admin/supplier/add', methods=['POST'])
@admin_login_required
def admin_supplier_add():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    company_name = request.form.get('company_name', '').strip()
    if not username or not password or not company_name:
        return redirect(url_for('admin_supplier_manage'))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            'INSERT INTO suppliers (username, password, company_name) VALUES (?, ?, ?)',
            (username, password, company_name)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.rollback()
    cursor.close()
    conn.close()
    return redirect(url_for('admin_supplier_manage'))

@app.route('/admin/supplier/delete/<int:supplier_id>', methods=['POST'])
@admin_login_required
def admin_supplier_delete(supplier_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM shipments WHERE supplier_id = ?', (supplier_id,))
    cursor.execute('DELETE FROM suppliers WHERE id = ?', (supplier_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('admin_supplier_manage'))

@app.route('/admin/delete/<int:item_id>', methods=['POST'])
@admin_login_required
def admin_delete(item_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM shipments WHERE id = ?', (item_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export/excel')
@admin_login_required
def admin_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return admin_export_csv()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT s.*, sp.company_name
        FROM shipments s
        JOIN suppliers sp ON s.supplier_id = sp.id
        ORDER BY sp.company_name, s.created_at DESC
    ''')
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '全部发货记录'

    headers = ['供应商', '项目名称', '项目编号', '采购订单号', '客户订单号',
                '物流单号', '物流回单', '发货日期', '备注', '提交时间']
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='E60012', end_color='E60012', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        values = [row['company_name'], row['project_name'], row['project_code'],
                  row['purchase_order'], row['customer_order'],
                  row['logistics_no'], row['logistics_receipt'],
                  str(row['ship_date']) if row['ship_date'] else '',
                  row['remark'] or '',
                  str(row['created_at']) if row['created_at'] else '']
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                      for r in range(1, len(rows) + 2))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=全批发货记录.xlsx'
    return response

# ==================== 初始化 ====================

@app.route('/init')
def init_route():
    init_db()
    admin_user = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_pass = os.environ.get('ADMIN_PASSWORD', 'admin123')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM admins')
    if cursor.fetchone()[0] == 0:
        cursor.execute(
            'INSERT INTO admins (username, password) VALUES (?, ?)',
            (admin_user, admin_pass)
        )
        conn.commit()
    cursor.close()
    conn.close()
    return '数据库初始化完成！<a href="/">返回首页</a>'

# PythonAnywhere WSGI 入口
application = app

if __name__ == '__main__':
    app.run(debug=True)
